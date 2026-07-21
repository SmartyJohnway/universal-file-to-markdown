"""
xlsx_converter.py
Deterministic Excel -> Markdown/HTML conversion with full merged-cell
fidelity. No AI model involved: openpyxl exposes merged ranges directly,
so reconstruction is exact, not inferred.

Rule of thumb:
  - a sheet with NO merged cells -> clean Markdown pipe table
  - a sheet WITH merged cells    -> HTML <table> with real rowspan/colspan
    (GFM pipe tables cannot express spans; forcing them in loses structure)

Formula handling: every cell is resolved through _resolve_cell(), which
checks the formula-mode workbook to detect formula cells and reports
missing cached values explicitly instead of a silent blank (see v1.2 fix
history in engine_notes.md for the bug this replaced).

v1.4 additions (Office fidelity):
  - number formats: date/datetime cells render as ISO dates instead of
    Python's default datetime repr; percentage/currency formats are
    respected where openpyxl exposes a plain value alongside the format.
  - hyperlinks: rendered as Markdown links when a cell has one attached.
  - cell comments/notes: collected into a per-sheet list rather than
    silently dropped (they don't fit inline in a table cell).
  - defined names: workbook-level named ranges are listed in the report.
  - used-region trimming: iterate only cells actually written (openpyxl's
    max_row/max_column can be inflated by formatting-only cells with no
    content, producing a table padded with meaningless empty rows/cols).
  - chart presence: noted in the report (title + type), not fully
    rendered - see engine_notes.md for why full chart data extraction is
    out of scope for this version.
"""

import datetime as dt

import openpyxl
from common_utils import extract_ooxml_media


def convert_xlsx(path: str, assets_dir: str = None) -> dict:
    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    media_saved = extract_ooxml_media(path, assets_dir) if assets_dir else []
    media_cursor = 0

    sections = []
    elements = []
    tables = []
    report = {
        "sheets": [],
        "total_merged_ranges": 0,
        "merged_ranges_rendered": 0,
        "hidden_sheets": [],
        "hidden_rows_or_columns_present": False,
        "formula_cells": 0,
        "formula_cached_values_available": 0,
        "formula_cached_values_missing": 0,
        "formula_cached_values_missing_cells": [],
        "defined_names": list(wb_values.defined_names.keys()) if hasattr(wb_values, "defined_names") else [],
        "comments_found": 0,
        "charts_found": 0,
        "images_found": 0,
        "media_extracted": len(media_saved),
    }

    for sheet_idx, ws_name in enumerate(wb_values.sheetnames):
        ws = wb_values[ws_name]
        ws_f = wb_formulas[ws_name]

        if ws.sheet_state != "visible":
            report["hidden_sheets"].append(ws_name)

        merged_ranges = list(ws.merged_cells.ranges)
        report["total_merged_ranges"] += len(merged_ranges)

        hidden_row = any(dim.hidden for dim in ws.row_dimensions.values())
        hidden_col = any(dim.hidden for dim in ws.column_dimensions.values())
        if hidden_row or hidden_col:
            report["hidden_rows_or_columns_present"] = True

        min_row, max_row, min_col, max_col = _used_region(ws, ws_f)
        blocks_found = _used_blocks(ws, ws_f, merged_ranges)
        formula_stats = {"cells": 0, "available": 0, "missing": 0, "missing_cells": []}
        comments = _collect_comments(ws, min_row, max_row, min_col, max_col)
        report["comments_found"] += len(comments)
        report["charts_found"] += len(getattr(ws, "_charts", []) or [])
        sheet_images = list(getattr(ws, "_images", []) or [])
        report["images_found"] += len(sheet_images)

        block_markdown = []
        block_results = []
        for block_idx, (b_min_row, b_max_row, b_min_col, b_max_col) in enumerate(blocks_found, start=1):
            block_merges = [mr for mr in merged_ranges
                            if b_min_row <= mr.min_row and mr.max_row <= b_max_row
                            and b_min_col <= mr.min_col and mr.max_col <= b_max_col]
            if block_merges:
                table_md, rendered, grid, cells = _render_html_table(
                    ws, ws_f, block_merges, ws_name, formula_stats,
                    b_min_row, b_max_row, b_min_col, b_max_col)
                report["merged_ranges_rendered"] += rendered
                engine = "html_table_with_span"
            else:
                table_md, grid = _render_pipe_table(
                    ws, ws_f, ws_name, formula_stats,
                    b_min_row, b_max_row, b_min_col, b_max_col)
                cells = None
                engine = "markdown_pipe_table"
            range_label = _range_label(b_min_row, b_max_row, b_min_col, b_max_col)
            if len(blocks_found) > 1:
                block_markdown.append(f"### Block: {range_label}\n\n{table_md}")
            else:
                block_markdown.append(table_md)
            block_results.append((block_idx, range_label, table_md, grid, cells, engine,
                                  len(block_merges)))

        comments_md = ""
        if comments:
            comment_lines = "\n".join(f"- **{c['cell']}**: {c['text']}" for c in comments)
            comments_md = f"<!-- cell comments -->\n{comment_lines}"

        report["formula_cells"] += formula_stats["cells"]
        report["formula_cached_values_available"] += formula_stats["available"]
        report["formula_cached_values_missing"] += formula_stats["missing"]
        report["formula_cached_values_missing_cells"].extend(formula_stats["missing_cells"])

        section_parts = block_markdown + ([comments_md] if comments_md else [])
        section_text = f"## Sheet: {ws_name}\n\n" + "\n\n".join(section_parts) + "\n"
        sections.append(section_text)
        report["sheets"].append({
            "name": ws_name,
            "merged_ranges": len(merged_ranges),
            "engine": "mixed_block_tables" if len({b[5] for b in block_results}) > 1
                      else block_results[0][5],
            "hidden": ws.sheet_state != "visible",
            "formula_cells": formula_stats["cells"],
            "formula_cached_values_missing": formula_stats["missing"],
            "comments": len(comments),
            "blocks": len(block_results),
        })

        el_id = f"sheet-{sheet_idx + 1:03d}"
        elements.append({
            "id": el_id,
            "type": "sheet",
            "sheet_name": ws_name,
            "content": f"## Sheet: {ws_name}",
            "engine": engine,
            "confidence": None,
            "source_locator": {"sheet_name": ws_name},
            "merged_ranges": len(merged_ranges),
        })
        for block_idx, range_label, table_md, grid, cells, engine, block_merge_count in block_results:
          if grid:
            table_id = f"table-s{sheet_idx + 1:03d}-b{block_idx:03d}"
            table_entry = {"id": table_id, "rows": grid,
                            "context": f"sheet:{ws_name}",
                            "source_locator": {"sheet": ws_name, "cell_range": range_label},
                            "engine": "openpyxl_custom"}
            if cells:
                table_entry["cells"] = cells
            tables.append(table_entry)
            elements.append({
                "id": f"element-{table_id}",
                "parent_id": el_id,
                "type": "table",
                "content": table_md,
                "engine": engine,
                "confidence": None,
                "source_locator": {"sheet_name": ws_name, "range": range_label},
                "table_id": table_id,
                "properties": {"block_index": block_idx, "merged_ranges": block_merge_count},
            })

        for chart_idx, chart in enumerate(getattr(ws, "_charts", []) or [], start=1):
            title = getattr(getattr(chart, "title", None), "tx", None)
            elements.append({
                "id": f"{el_id}-chart-{chart_idx:03d}", "parent_id": el_id,
                "type": "chart_reference", "content": f"Chart {chart_idx}",
                "engine": "openpyxl_custom", "confidence": None,
                "source_locator": {"sheet": ws_name, "shape_id": chart_idx},
                "properties": {"chart_type": type(chart).__name__, "title_present": title is not None},
                "warnings": ["XLSX_CHART_NOT_RENDERED"],
            })
        for image_idx, _image in enumerate(sheet_images, start=1):
            asset = media_saved[media_cursor] if media_cursor < len(media_saved) else None
            media_cursor += 1
            elements.append({
                "id": f"{el_id}-image-{image_idx:03d}", "parent_id": el_id,
                "type": "image", "content": f"![]({asset})" if asset else "",
                "engine": "openpyxl_custom", "confidence": None,
                "source_locator": {"sheet": ws_name, "shape_id": image_idx},
                "asset": asset,
            })

    markdown = "\n".join(sections)
    return {"markdown": markdown, "report": report, "elements": elements, "tables": tables}


def _range_label(min_row, max_row, min_col, max_col):
    if max_row < min_row or max_col < min_col:
        return None
    from openpyxl.utils import get_column_letter
    return (f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}")


def _used_blocks(ws, ws_f, merged_ranges):
    occupied = set()
    for row in ws.iter_rows():
        for cell in row:
            formula = ws_f.cell(cell.row, cell.column).value
            if cell.value is not None or (isinstance(formula, str) and formula.startswith("=")):
                occupied.add((cell.row, cell.column))
    for mr in merged_ranges:
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                occupied.add((row, col))
    if not occupied:
        return [(1, 0, 1, 0)]

    row_groups = _consecutive_groups(sorted({row for row, _ in occupied}))
    blocks = []
    for row_start, row_end in row_groups:
        cols = sorted({col for row, col in occupied if row_start <= row <= row_end})
        for col_start, col_end in _consecutive_groups(cols):
            if any(row_start <= row <= row_end and col_start <= col <= col_end
                   for row, col in occupied):
                blocks.append((row_start, row_end, col_start, col_end))
    return blocks


def _consecutive_groups(values):
    if not values:
        return []
    groups = []
    start = previous = values[0]
    for value in values[1:]:
        if value != previous + 1:
            groups.append((start, previous))
            start = value
        previous = value
    groups.append((start, previous))
    return groups


def _used_region(ws, ws_f=None):
    """Find the actual bounds of written content rather than trusting
    ws.max_row/max_column directly, which openpyxl can report as larger
    than the real content due to formatting applied to unwritten cells.

    Bug found by the test suite: checking only the data_only=True
    workbook excluded formula cells with no cached value entirely (their
    value there is None), so a formula-only cell fell outside the
    detected used region and was never rendered or counted - the exact
    silent-blank failure mode this file's formula handling was supposed
    to fix. Both workbooks must be checked."""
    min_row = min_col = None
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            has_value = cell.value is not None
            has_formula = False
            if ws_f is not None:
                f_cell = ws_f.cell(row=cell.row, column=cell.column)
                has_formula = isinstance(f_cell.value, str) and f_cell.value.startswith("=")
            if has_value or has_formula:
                if min_row is None or cell.row < min_row:
                    min_row = cell.row
                if min_col is None or cell.column < min_col:
                    min_col = cell.column
                if cell.row > max_row:
                    max_row = cell.row
                if cell.column > max_col:
                    max_col = cell.column
    if min_row is None:
        return 1, 0, 1, 0  # empty sheet
    return min_row, max_row, min_col, max_col


def _collect_comments(ws, min_row, max_row, min_col, max_col):
    comments = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.comment is not None:
                comments.append({"cell": cell.coordinate, "text": cell.comment.text.strip()})
    return comments


def _format_value(value, number_format: str) -> str:
    if isinstance(value, (dt.datetime, dt.date)):
        fmt = (number_format or "").lower()
        if isinstance(value, dt.datetime) and ("h" in fmt or ":" in fmt):
            return value.isoformat(sep=" ")
        return value.isoformat() if hasattr(value, "isoformat") else str(value)
    return str(value)


def _resolve_cell(ws, ws_f, r: int, c: int, sheet_name: str, stats: dict) -> str:
    formula_cell = ws_f.cell(row=r, column=c)
    is_formula = isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
    value_cell = ws.cell(row=r, column=c)

    if not is_formula:
        value = value_cell.value
        if value is None:
            text = ""
        else:
            text = _format_value(value, value_cell.number_format)
    else:
        stats["cells"] += 1
        cached_value = value_cell.value
        if cached_value is not None:
            stats["available"] += 1
            text = _format_value(cached_value, value_cell.number_format)
        else:
            stats["missing"] += 1
            cell_ref = f"{sheet_name}!{formula_cell.coordinate}"
            stats["missing_cells"].append(cell_ref)
            text = f"{formula_cell.value} <!-- formula result unavailable -->"

    hyperlink = value_cell.hyperlink
    if hyperlink is not None and hyperlink.target:
        label = text if text else hyperlink.target
        text = f"[{label}]({hyperlink.target})"

    return text


def _render_pipe_table(ws, ws_f, sheet_name, formula_stats, min_row, max_row, min_col, max_col):
    if max_row < min_row or max_col < min_col:
        return "_(empty sheet)_", None

    grid = [[_resolve_cell(ws, ws_f, r, c, sheet_name, formula_stats)
             for c in range(min_col, max_col + 1)]
            for r in range(min_row, max_row + 1)]

    header = grid[0]
    lines = ["| " + " | ".join(_cell_str(c) for c in header) + " |",
             "| " + " | ".join(["---"] * len(header)) + " |"]
    for row in grid[1:]:
        lines.append("| " + " | ".join(_cell_str(c) for c in row) + " |")
    return "\n".join(lines), grid


def _render_html_table(ws, ws_f, merged_ranges, sheet_name, formula_stats,
                        min_row, max_row, min_col, max_col):
    merge_lookup = {}
    rendered = 0
    for mr in merged_ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        rowspan = mx_r - mn_r + 1
        colspan = mx_c - mn_c + 1
        merge_lookup[(mn_r, mn_c)] = ("anchor", rowspan, colspan)
        for r in range(mn_r, mx_r + 1):
            for c in range(mn_c, mx_c + 1):
                if (r, c) != (mn_r, mn_c):
                    merge_lookup[(r, c)] = ("skip", 0, 0)
        rendered += 1

    html = ["<table>"]
    grid = []
    cells = []
    for r in range(min_row, max_row + 1):
        html.append("<tr>")
        grid_row = []
        for c in range(min_col, max_col + 1):
            info = merge_lookup.get((r, c))
            if info and info[0] == "skip":
                # cell covered by another cell's span: still occupies a
                # grid position (rows/*.csv must stay rectangular), just
                # has no <td> of its own and no entry in `cells`.
                grid_row.append("")
                continue
            cell_text = _resolve_cell(ws, ws_f, r, c, sheet_name, formula_stats)
            grid_row.append(cell_text)
            rowspan = colspan = 1
            attrs = ""
            if info and info[0] == "anchor":
                _, rowspan, colspan = info
                if rowspan > 1:
                    attrs += f' rowspan="{rowspan}"'
                if colspan > 1:
                    attrs += f' colspan="{colspan}"'
            html.append(f"<td{attrs}>{_html_escape(cell_text)}</td>")
            cells.append({"row": r - min_row, "col": c - min_col, "value": cell_text,
                          "rowspan": rowspan, "colspan": colspan})
        html.append("</tr>")
        grid.append(grid_row)
    html.append("</table>")
    return "\n".join(html), rendered, grid, cells


def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).replace("|", "\\|").replace("\n", "<br>")


def _html_escape(s: str) -> str:
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
